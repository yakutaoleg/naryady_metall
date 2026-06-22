from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "MVP Pricing"

BLUE    = PatternFill("solid", start_color="1F4E79")
GREEN   = PatternFill("solid", start_color="E2EFDA")
YELLOW  = PatternFill("solid", start_color="FFF2CC")
GRAY    = PatternFill("solid", start_color="D9D9D9")
WHITE   = PatternFill("solid", start_color="FFFFFF")

thin = Side(style="thin", color="BFBFBF")
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def align(h="left", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical="top", wrap_text=wrap, indent=indent)

# Column widths
for col, w in zip("ABCDEFG", [5, 26, 46, 16, 13, 13, 16]):
    ws.column_dimensions[col].width = w

# Title
ws.merge_cells("A1:G1")
c = ws["A1"]
c.value = "MVP — Система управления цехом металлоконструкций"
c.font  = font(bold=True, color="1F4E79", size=13)
c.alignment = align()
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:G2")
c = ws["A2"]
c.value = "Подготовлено: Олег Якута · 2026 · Конфиденциально"
c.font  = font(color="808080", size=9, italic=True)
c.alignment = align()
ws.row_dimensions[2].height = 16

# Header
ws.append([])
ws.row_dimensions[3].height = 4

headers = ["#", "Фича", "Что это значит", "Статус", "$", "BYN", "Выполнено"]
ws.append(headers)
for col, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=col)
    c.value     = h
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill      = BLUE
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = brd
ws.row_dimensions[4].height = 22

ROW = [4]

def section(title):
    ROW[0] += 1
    ws.merge_cells(f"A{ROW[0]}:G{ROW[0]}")
    c = ws[f"A{ROW[0]}"]
    c.value     = title
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill      = BLUE
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border    = brd
    ws.row_dimensions[ROW[0]].height = 18

def row(num, name, desc, status, usd, byn, done):
    ROW[0] += 1
    fill = GREEN if done == "выполнено" else (YELLOW if done == "в работе" else WHITE)
    vals = [num, name, desc, status, usd, byn, done]
    for col, v in enumerate(vals, 1):
        c = ws.cell(row=ROW[0], column=col)
        c.value  = v
        c.fill   = fill
        c.border = brd
        c.alignment = align(h="center" if col in (1, 4, 5, 6, 7) else "left",
                            wrap=True)
        if col in (5, 6) and v == "бесплатно":
            c.font = font(color="375623")
        elif col == 7 and v == "выполнено":
            c.font = font(color="375623")
        elif col == 7 and v == "в работе":
            c.font = font(color="7F6000")
        else:
            c.font = font()
    ws.row_dimensions[ROW[0]].height = 42

section("Фундамент")
row(1, "База данных и синхронизация",
    "Вся информация о нарядах, рабочих и статусах хранится в одном месте и автоматически обновляется.",
    "было", "$200", "570 BYN", "выполнено")
row(2, "История действий",
    "Каждое действие рабочего фиксируется: кто, что и когда сделал. Всегда можно разобрать спорную ситуацию.",
    "+ добавлено", "бесплатно", "бесплатно", "выполнено")

section("Бот для рабочих")
row(3, "Список задач и статусы",
    "Рабочий открывает Telegram и видит только свои задачи в нужном порядке. Отмечает выполнено или сигнализирует о проблеме.",
    "было", "$200", "570 BYN", "выполнено")
row(4, "Сумма к оплате",
    "Рабочий сразу видит сколько заработает за каждую позицию. Прозрачно для всех сторон.",
    "+ добавлено", "бесплатно", "бесплатно", "выполнено")
row(5, "Чертёж прямо в чат",
    "Рабочий нажимает кнопку — бот присылает чертёж на телефон. Не нужно идти в офис или звонить мастеру.",
    "+ добавлено", "$100", "285 BYN", "выполнено")
row(6, "Онбординг рабочих",
    "Регистрация каждого рабочего в системе, первый вход, проверка что всё работает корректно.",
    "+ добавлено", "бесплатно", "бесплатно", "запланировано")

section("Блокировки и уведомления")
row(7, "Мгновенные оповещения о проблемах",
    "Если рабочий не может выполнить позицию — указывает причину, мастер сразу получает уведомление. Проблемы не замалчиваются.",
    "было", "$150", "428 BYN", "запланировано")

section("Логика производства")
row(8, "Зависимости между этапами",
    "Сборщик не получит задачу пока пила и плазма не закроют все детали по элементу. Система контролирует готовность автоматически.",
    "было", "$150", "428 BYN", "запланировано")

section("Выработка")
row(9, "Автоматический расчёт заработка",
    "Система считает выработку каждого рабочего по мере выполнения задач. Руководство видит кто сколько сделал без ручного подсчёта.",
    "было", "$100", "285 BYN", "запланировано")

section("Автозаполнение нарядов из документации проектировщика")
row(10, "Генерация нарядов из файлов КМД",
    "Проектировщик передаёт технические файлы — система сама разбивает их на наряды по специализациям и формирует готовый документ. Ручной ввод исчезает полностью.",
    "+ добавлено", "$350", "998 BYN", "в работе")

section("Запуск")
row(11, "Сервер и деплой",
    "Система разворачивается на сервере, настраивается тестовая и боевая версии. Клиент ничего не настраивает сам.",
    "+ добавлено", "бесплатно", "бесплатно", "выполнено")
row(12, "Обучение мастера и администратора",
    "Разбираем как загружать наряды, расставлять приоритеты и управлять системой. Остаёмся на связи пока всё не встанет как надо.",
    "+ добавлено", "бесплатно", "бесплатно", "запланировано")

def total(label, usd, byn):
    ROW[0] += 1
    ws.merge_cells(f"A{ROW[0]}:D{ROW[0]}")
    c = ws[f"A{ROW[0]}"]
    c.value     = label
    c.font      = Font(name="Arial", bold=True, size=10)
    c.fill      = GRAY
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.border    = brd
    for col, v in [(5, usd), (6, byn), (7, "")]:
        c = ws.cell(row=ROW[0], column=col)
        c.value     = v
        c.font      = Font(name="Arial", bold=True, size=10)
        c.fill      = GRAY
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = brd
    ws.row_dimensions[ROW[0]].height = 20

total("MVP (без КМД)", "$1 500", "4 275 BYN")
total("MVP + автозаполнение нарядов из КМД", "$1 850", "5 273 BYN")

# Footer
ROW[0] += 2
ws.merge_cells(f"A{ROW[0]}:G{ROW[0]}")
c = ws[f"A{ROW[0]}"]
c.value     = "Курс: 1$ = 2.85 BYN"
c.font      = Font(name="Arial", size=9, color="808080", italic=True)
c.alignment = Alignment(horizontal="right")

out = "C:/Users/MSI/Projects/Automatization/Metall/docs/MVP_pricing.xlsx"
wb.save(out)
print("OK:", out)
